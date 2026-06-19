import logging
from decimal import Decimal
from io import BytesIO

from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.crud.almacen import crud_almacen
from app.events.publisher import publish_event
from app.models.almacen import Almacen
from app.models.existencia import Existencia
from app.schemas.stock import StockIngresoRequest
from app.services.catalogos import catalogos_client
from app.services.stock import stock_service

logger = logging.getLogger(__name__)

COLUMNAS = ("codigo", "producto", "sucursal", "cantidad", "costo", "precio")


class ExcelInventarioService:
    async def _resolver_almacen(self, db: AsyncSession, sucursal_nombre: str) -> Almacen:
        nombre = sucursal_nombre.strip()
        stmt = select(Almacen).where(
            (Almacen.nombre.ilike(f"%{nombre}%")) | (Almacen.codigo.ilike(f"%{nombre}%"))
        )
        almacen = (await db.execute(stmt)).scalar_one_or_none()
        if almacen:
            return almacen

        codigo = f"ALM-{nombre.upper().replace(' ', '-')[:20]}"
        almacen = Almacen(codigo=codigo, nombre=f"Almacén {nombre}", direccion=nombre)
        db.add(almacen)
        await db.flush()
        return almacen

    async def cargar_excel(self, db: AsyncSession, file: UploadFile) -> dict:
        if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
            raise HTTPException(status_code=400, detail="Archivo debe ser Excel (.xlsx)")

        content = await file.read()
        wb = load_workbook(BytesIO(content), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(status_code=400, detail="Excel vacío")

        headers = [str(h).strip().lower() if h else "" for h in rows[0]]
        for col in COLUMNAS:
            if col not in headers:
                raise HTTPException(status_code=400, detail=f"Columna requerida faltante: {col}")

        idx = {col: headers.index(col) for col in COLUMNAS}
        procesados = 0
        errores: list[str] = []

        for i, row in enumerate(rows[1:], start=2):
            if not row or not row[idx["codigo"]]:
                continue
            try:
                codigo = str(row[idx["codigo"]]).strip()
                sucursal = str(row[idx["sucursal"]]).strip()
                cantidad = Decimal(str(row[idx["cantidad"]]))

                producto = await catalogos_client.obtener_por_codigo(codigo)
                almacen = await self._resolver_almacen(db, sucursal)

                await stock_service.ingreso(
                    db,
                    StockIngresoRequest(
                        producto_id=producto["id"],
                        almacen_id=almacen.id,
                        cantidad=cantidad,
                        observaciones=f"Carga Excel: {producto.get('nombre', codigo)}",
                    ),
                )
                procesados += 1
            except Exception as exc:
                errores.append(f"Fila {i}: {exc}")
                logger.warning("Error fila %s: %s", i, exc)

        await publish_event("InventoryLoaded", {"filas_procesadas": procesados, "errores": len(errores)})
        return {"procesados": procesados, "errores": errores}


excel_service = ExcelInventarioService()


async def saldo_consolidado(db: AsyncSession, producto_id: int) -> dict:
    stmt = (
        select(
            Existencia.producto_id,
            func.sum(Existencia.cantidad_actual).label("cantidad_total"),
            func.count(Existencia.id).label("almacenes"),
        )
        .where(Existencia.producto_id == producto_id)
        .group_by(Existencia.producto_id)
    )
    row = (await db.execute(stmt)).one_or_none()
    if not row:
        return {"producto_id": producto_id, "cantidad_total": 0, "almacenes": 0}
    return {
        "producto_id": producto_id,
        "cantidad_total": float(row.cantidad_total or 0),
        "almacenes": int(row.almacenes or 0),
    }
